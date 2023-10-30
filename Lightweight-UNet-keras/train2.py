import glob
import numpy as np
import tensorflow as tf
# import unet
# import U2NET
# import LUNet
import DeepLabV3p
import matplotlib.pyplot as plt
from keras.callbacks import LearningRateScheduler
from keras.callbacks import EarlyStopping, ModelCheckpoint
import openpyxl


images = glob.glob('../DATA/2labels/train_crop/*image.png')
annotations = glob.glob('../DATA/2labels/train_crop/*label.png')
# index = np.random.permutation(len(images))
images = np.array(images)
annos = np.array(annotations)
dataset = tf.data.Dataset.from_tensor_slices((images, annos))
# 后20% 是测试集
test_count = int(len(images) *0.1)
train_count = len(images) - test_count
dataset_train = dataset.take(train_count)
dataset_test = dataset.skip(train_count)

# image_size = (304, 480)
image_size = (320, 384)
num_classes = 3

# weights_name = 'LUNet3a_crop_again.h5'
# weights_name = 'UNET_crop.h5'
weights_name = 'DeepLabV3-resnet50.h5'
BATCH_SIZE = 8

hist_save_path = "./test_imgs/2labels/"+weights_name[:-3]

# class MeanIoU(tf.keras.metrics.MeanIoU): 
    # def update_state(self, y_true, y_pred, sample_weight=None):
        # y_pred = tf.argmax(y_pred, axis=-1)
        # return super().update_state(y_true, y_pred, sample_weight=sample_weight)


class MeanIoU(tf.keras.metrics.MeanIoU):

    def update_state(self, y_true, y_pred, sample_weight=None):
        y_true = tf.argmax(y_true, axis=-1)
        # y_true = tf.keras.utils.to_categorical(y_true, num_classes=2)
        y_pred = tf.argmax(y_pred, axis=-1)
        super().update_state(y_true, y_pred, sample_weight=sample_weight)
        return self.result()


def read_png(path, channels):
    img = tf.io.read_file(path)
    img = tf.image.decode_png(img, channels=channels)
    return img

def normalize(input_image, input_mask):
    input_image = tf.cast(input_image, tf.float32)/255.0
    return input_image, input_mask

def load_image(input_image_path, input_mask_path):
    input_image = read_png(input_image_path, 3)
    input_mask = read_png(input_mask_path, 1)
    # 587 945 ->  304  480
    input_image = tf.image.resize(input_image, image_size)
    input_mask = tf.image.resize(input_mask, image_size)
    # if tf.random.uniform(()) > 0.5:
        # input_image = tf.image.flip_left_right(input_image)
        # input_mask = tf.image.flip_left_right(input_mask)
    input_image, input_mask = normalize(input_image, input_mask)
    return input_image, input_mask


STEPS_PER_EPOCH = int(train_count / BATCH_SIZE)-1
dataset_train = dataset_train.map(load_image, num_parallel_calls=tf.data.experimental.AUTOTUNE)
dataset_train = dataset_train.shuffle(train_count).batch(BATCH_SIZE)
# dataset_train = dataset_train.batch(BATCH_SIZE).repeat()
dataset_train = dataset_train.prefetch(buffer_size=tf.data.experimental.AUTOTUNE)

dataset_test = dataset_test.map(load_image, num_parallel_calls=tf.data.experimental.AUTOTUNE)
dataset_test = dataset_test.batch(BATCH_SIZE)

  
# if(1):
#     model = unet.build_model((image_size[0], image_size[1], 3), 8)
# else:
#     model = tf.keras.models.load_model('./UNET--.h5')

# model = unet.build_model((image_size[0], image_size[1], 3), num_classes)
# model = U2NET.u2net(image_size, out_ch=num_classes)
# model = LUNet.build_model((image_size[0], image_size[1], 3), num_classes)
# model = deeplabv3.Deeplabv3(input_shape=(image_size[0], image_size[1], 3), classes=num_classes, backbone='mobilenetv2',activation='sigmoid')
model = DeepLabV3p.DeeplabV3Plus((image_size[0], image_size[1], 3), num_classes)
# model = tf.keras.models.load_model("./UNET2.h5")

checkpointer = ModelCheckpoint('./weights/2labels/'+weights_name, monitor=MeanIoU(num_classes=num_classes),verbose=1, save_best_only=True,mode='max')

def display(display_list):
    plt.figure(figsize=(15, 5))
    title = ['Input Image', 'True Mask', 'Predicted Mask']
    for i in range(len(display_list)):
        plt.subplot(1, len(display_list), i+1)
        plt.title(title[i])
        plt.imshow(tf.keras.preprocessing.image.array_to_img(display_list[i]))
    plt.show()


def create_mask(pred_mask):
    pred_mask = tf.argmax(pred_mask, axis=-1)
    pred_mask = pred_mask[..., tf.newaxis]
    return pred_mask[0]


def show_predictions():
    for image, mask in dataset_train.take(1):
        pred_mask = model.predict(image)
        display([image[0], mask[0], create_mask(pred_mask)])

def save_data_xlsx(p,data):
    row = 1
    for xx in data:
        outws.cell(column = p , row = row).value = xx
        row += 1
    outwb.save(hist_save_path+'.xlsx')  # 保存结果

# class DisplayCallback(tf.keras.callbacks.Callback):
    # def on_epoch_end(self, epoch, logs=None):
        # pass
        # show_predictions()


model.summary()
adam = tf.keras.optimizers.Adam(lr=0.001, beta_1=0.9, beta_2=0.999, epsilon=None, decay=0.0, amsgrad=False)
# model.compile(optimizer='adam', loss='sparse_categorical_crossentropy', metrics=['acc', MeanIoU(num_classes=num_classes)])
model.compile(optimizer='adam', loss='sparse_categorical_crossentropy', metrics=['accuracy', MeanIoU(num_classes=num_classes)])

EPOCHS = 100
tf.config.experimental_run_functions_eagerly(True)
# history = model.fit(dataset_train, 
                    # epochs=EPOCHS,
                    # steps_per_epoch=STEPS_PER_EPOCH,
                    # callbacks=[DisplayCallback()])

### Additions ###
# 自动调整lr值 #
def lr_scheduler(epoch, lr):
    decay_rate = 0.85
    decay_step = 30
    if epoch % decay_step == 0 and epoch:
        return lr * pow(decay_rate, np.floor(epoch / decay_step))
    return lr
### End ###

callbacks = [LearningRateScheduler(lr_scheduler, verbose=1)]
history = model.fit(dataset_train,
                    callbacks=[checkpointer],
                    epochs=EPOCHS,
                    validation_data=dataset_test)

### Print history ###
loss = np.array(history.history['loss'])
val_loss = np.array(history.history['val_loss'])
acc = np.array(history.history['accuracy'])
val_acc = np.array(history.history['val_accuracy'])
x_scope = np.linspace(1,EPOCHS,num=len(loss))

plt.plot(x_scope, loss)
plt.plot(x_scope, val_loss)
plt.plot(x_scope, acc)
plt.plot(x_scope, val_acc)
plt.title("model loss & accuracy")
plt.ylabel("acc-loss")
plt.xlabel("epoch")
plt.legend(["train","test"],loc="upper left")
plt.show()

# Save as xlsx #
outwb = openpyxl.Workbook()  # 打开一个将写的文件
outws = outwb.create_sheet(index=0)  # 在将写的文件创建sheet

save_data_xlsx(1,loss)
save_data_xlsx(2,acc)
save_data_xlsx(3,val_loss)
save_data_xlsx(4,val_acc)
# End #
### End ###

model.save('./weights/2labels/'+weights_name, include_optimizer = False)

